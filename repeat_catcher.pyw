from openpyxl import load_workbook
from tkinter import *
from tkinter import messagebox


def save():
    xl = 'test.xlsx'
    file = load_workbook(xl)
    sheet = file['Лист1']
    col = int(e.get())
    data = sheet.values
    data = list(data)
    first_column = [r[0] for r in data]
    for row, item in enumerate(first_column):
        value = first_column.count(item)
        coord = sheet.cell(row=row + 1, column=col)
        coord.value = value
    file.save(xl)
    file.close()
    messagebox.askokcancel('Сохранение', 'Успешно!')


root = Tk()
root.title('Repeat catcher v1.0')
root.geometry('500x200')
root.resizable(0, 0)


lb = Label(root, text='Введите номер пустой колонки, например: 24', font='Arial 8')
lb.pack()
e = Entry(root)
e.pack()
root.title('Нажмите на обработку!')
btn = Button(root, text='Обработать!', font='Arial 10', command=save)
btn.pack()

if __name__ == '__main__':
    root.mainloop()

